from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
from dp_excel.ExcelFile import ExcelFile


class ExcelCellOptions:
    OPTION_COLSPAN = 'colspan'
    OPTION_ROWSPAN = 'rowspan'
    OPTION_FONT = 'font'
    OPTION_FILL = 'fill'
    OPTION_BORDER = 'border'
    OPTION_ALIGNMENT = 'alignment'
    OPTION_NUMBER_FORMAT = 'number_format'
    OPTION_PROTECTION = 'protection'
    OPTION_WIDTH = 'width'

    def __init__(self):
        self.options = []

    def _add_option(self, option, value):
        self.options.append((option, value))
        return self

    def get_options(self):
        for option in self.options:
            yield option

    def set_colspan(self, count):
        self._add_option(self.OPTION_COLSPAN, count)
        return self

    def set_rowspan(self, count):
        self._add_option(self.OPTION_ROWSPAN, count)
        return self

    def set_font(self, name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none',
                 strike=False, color='FF000000'):
        font = Font(name=name, size=size, bold=bold, italic=italic, vertAlign=vertAlign, underline=underline,
                    strike=strike, color=color)
        self._add_option(self.OPTION_FONT, font)
        return self

    def set_fill(self, fill_type=None, start_color='FFFFFFFF', end_color='FF000000'):
        fill = PatternFill(fill_type='solid', start_color=start_color, end_color=end_color)
        self._add_option(self.OPTION_FILL, fill)
        return self

    def set_border(self, left=Side(border_style=None, color='FF000000'),
                   right=Side(border_style=None, color='FF000000'),
                   top=Side(border_style=None, color='FF000000'),
                   bottom=Side(border_style=None, color='FF000000'),
                   diagonal=Side(border_style=None, color='FF000000'),
                   diagonal_direction=0,
                   outline=Side(border_style=None, color='FF000000'),
                   vertical=Side(border_style=None, color='FF000000'),
                   horizontal=Side(border_style=None, color='FF000000'),
                   perimeter=Side(border_style=None, color='FF000000')):

        if perimeter:
            top = right = bottom = left = perimeter

        border = Border(left=left, right=right, top=top, bottom=bottom, diagonal=diagonal,
                        diagonal_direction=diagonal_direction, outline=outline, vertical=vertical,
                        horizontal=horizontal)
        self._add_option(self.OPTION_BORDER, border)
        return self

    def set_alignment(self, horizontal='general', vertical='bottom', text_rotation=0, wrap_text=False,
                      shrink_to_fit=False, indent=0):
        alignment = Alignment(horizontal=horizontal, vertical=vertical, text_rotation=text_rotation,
                              wrap_text=wrap_text, shrink_to_fit=shrink_to_fit, indent=indent)

        self._add_option(self.OPTION_ALIGNMENT, alignment)
        return self

    def set_number_format(self, number_format='General'):
        self._add_option(self.OPTION_NUMBER_FORMAT, number_format)
        return self

    def set_protection(self, locked=True, hidden=False):
        protection = Protection(locked=locked, hidden=hidden)
        self._add_option(self.OPTION_PROTECTION, protection)
        return self

    def set_width(self, width):
        self._add_option(self.OPTION_WIDTH, width)
        return self

    def apply_colspan(self, worksheet, value, column_number, row_number):
        from_idx = ExcelFile.get_idx(column_number, row_number)
        to_idx = ExcelFile.get_idx(column_number + value - 1, row_number)
        worksheet.merge_cells('{from_cell}:{to_cell}'.format(from_cell=from_idx, to_cell=to_idx))

    def apply_rowspan(self, worksheet, value, column_number, row_number):
        from_idx = ExcelFile.get_idx(column_number, row_number)
        to_idx = ExcelFile.get_idx(column_number, row_number + value - 1)
        worksheet.merge_cells('{from_cell}:{to_cell}'.format(from_cell=from_idx, to_cell=to_idx))

    def apply_font(self, worksheet, value, column_number, row_number):
        if isinstance(value, Font):
            worksheet[ExcelFile.get_idx(column_number, row_number)].font = value
        else:
            raise ValueError('font option must be is instance of openpyxl.styles.Font')

    def apply_fill(self, worksheet, value, column_number, row_number):
        if isinstance(value, PatternFill):
            worksheet[ExcelFile.get_idx(column_number, row_number)].fill = value
        else:
            raise ValueError('font option must be is instance of openpyxl.styles.PatternFill')

    def apply_border(self, worksheet, value, column_number, row_number):
        if isinstance(value, Border):
            worksheet[ExcelFile.get_idx(column_number, row_number)].border = value
        else:
            raise ValueError('font option must be is instance of openpyxl.styles.Border')

    def apply_alignment(self, worksheet, value, column_number, row_number):
        if isinstance(value, Alignment):
            worksheet[ExcelFile.get_idx(column_number, row_number)].alignment = value
        else:
            raise ValueError('font option must be is instance of openpyxl.styles.Alignment')

    def apply_number_format(self, worksheet, value, column_number, row_number):
        if isinstance(value, str):
            worksheet[ExcelFile.get_idx(column_number, row_number)].number_format = value
        else:
            raise ValueError('font option must be is instance of Str')

    def apply_protection(self, worksheet, value, column_number, row_number):
        if isinstance(value, Protection):
            worksheet[ExcelFile.get_idx(column_number, row_number)].protection = value
        else:
            raise ValueError('font option must be is instance of openpyxl.styles.Protection')

    def apply_width(self, worksheet, value, column_number):
        worksheet.column_dimensions[get_column_letter(column_number)].width = value

    def apply_cell_options(self, worksheet, column_number, row_number):
        for option, value in self.get_options():
            if option == self.OPTION_COLSPAN:
                self.apply_colspan(worksheet, value, column_number, row_number)
            if option == self.OPTION_ROWSPAN:
                self.apply_rowspan(worksheet, value, column_number, row_number)
            elif option == self.OPTION_FONT:
                self.apply_font(worksheet, value, column_number, row_number)
            elif option == self.OPTION_FILL:
                self.apply_fill(worksheet, value, column_number, row_number)
            elif option == self.OPTION_BORDER:
                self.apply_border(worksheet, value, column_number, row_number)
            elif option == self.OPTION_ALIGNMENT:
                self.apply_alignment(worksheet, value, column_number, row_number)
            elif option == self.OPTION_NUMBER_FORMAT:
                self.apply_number_format(worksheet, value, column_number, row_number)
            elif option == self.OPTION_PROTECTION:
                self.apply_protection(worksheet, value, column_number, row_number)
            elif option == self.OPTION_WIDTH:
                self.apply_width(worksheet, value, column_number)
