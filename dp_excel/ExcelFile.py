from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook


class ExcelFile:

    def __init__(self, file_name):
        self.file_name = file_name
        self.workbook = Workbook()
        self.active_worksheet = self.workbook.active
        self.current_position = {'col': 1, 'row': 1}

    @staticmethod
    def get_idx(column_number, row_number):
        return '{column}{row}'.format(column=get_column_letter(column_number), row=row_number)

    def add_worksheet(self, title=None, index=None, set_is_active=True):
        worksheet = self.workbook.create_sheet(title=title, index=index)

        if set_is_active:
            self.set_active_worksheet(worksheet)

    def get_worksheet_by_title(self, title):
        for sheet in self.workbook:
            if sheet.title == title:
                return sheet

        raise ValueError('Worksheet with title {0} Not Found'.format(title))

    def set_active_worksheet(self, worksheet):
        self.active_worksheet = worksheet

    def set_active_worksheet_by_title(self, title):
        self.active_worksheet = self.get_worksheet_by_title(title=title)

    def set_position(self, column_number, row_number):
        """
        Set active excel cell
        :param column_number: integer
        :param row_number:    integer
        :return:
        """
        self.current_position = {'col': column_number, 'row': row_number}
        return self

    def render_row_by_template(self, template):
        """
        Render row by template

        :param template: ExcelRowTemplate object
        :return:
        """
        start_col_idx = self.current_position['col']
        for row in template.get_rows():
            self.current_position['col'] = start_col_idx
            template.get_row_options(row).apply_row_options(self.active_worksheet, self.current_position['row'])

            for cell in template.get_columns(row):
                if not cell.is_empty:
                    col_idx = get_column_letter(self.current_position['col'])
                    row_idx = self.current_position['row']
                    cell_idx = '{column}{row}'.format(column=col_idx, row=row_idx)
                    self.active_worksheet[cell_idx] = cell.value
                    cell.options.apply_cell_options(self.active_worksheet, self.current_position['col'], row_idx)

                self.current_position['col'] += 1

            self.current_position['row'] += 1

        self.current_position['col'] = start_col_idx

    def save(self, path=None):
        self.workbook.save('{0}{1}.xlsx'.format(path if path else '', self.file_name))

    def get_virtual_workbook(self):
        """
        EXAMPLE USING return HttpResponse(excel_file.get_virtual_workbook(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        :return:
        """
        return save_virtual_workbook(self.workbook)
